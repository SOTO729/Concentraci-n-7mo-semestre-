import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Respuestas Ejercicios"

# Definir estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Encabezados
headers = ["Ejercicio", "Tema", "Instrucción", "Código Respuesta"]
ws.append(headers)

# Aplicar estilos al encabezado
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# Datos de los ejercicios
ejercicios = [
    {
        "numero": 1,
        "tema": "Variables de cadena (String)",
        "instruccion": "Crea una variable llamada 'despedida' que guarde 'Adiós mundo' e imprime la variable.",
        "codigo": "despedida='Adiós mundo'\nprint(despedida)"
    },
    {
        "numero": 2,
        "tema": "Variable mal utilizada",
        "instruccion": "Reemplaza la palabra reservada 'else' correctamente en el código.",
        "codigo": "x=5\nif x>0:\n    print('Positivo')\nelse:\n    print('No positivo')"
    },
    {
        "numero": 3,
        "tema": "Comentario en función",
        "instruccion": "Agrega el comentario 'Esta función calcula el producto de dos números.' a la función.",
        "codigo": "# Esta función calcula el producto de dos números.\ndef multiplicar(a,b):\n    return a*b"
    },
    {
        "numero": 4,
        "tema": "Variables enteras",
        "instruccion": "Crea variables contador1=10 y contador2=20 e imprime ambas con una sola función print.",
        "codigo": "contador1=10\ncontador2=20\nprint(contador1,contador2)"
    },
    {
        "numero": 5,
        "tema": "Variable flotante",
        "instruccion": "Crea una variable tipo flotante 'total' con valor 25500.75 e imprime.",
        "codigo": "total=25500.75\nprint(total)"
    },
    {
        "numero": 6,
        "tema": "Conversión de tipos",
        "instruccion": "Transforma 'n_libros' a cadena e imprime con 'mensaje' usando una sola función print.",
        "codigo": "n_libros=150\nmensaje='La biblioteca tiene los siguientes libros:'\nprint(mensaje,str(n_libros))"
    },
    {
        "numero": 7,
        "tema": "Variable tipo lista",
        "instruccion": "Crea una lista 'colores' con valores 'rojo','verde','azul' e imprime.",
        "codigo": "colores=['rojo','verde','azul']\nprint(colores)"
    },
    {
        "numero": 8,
        "tema": "Variable tipo tupla",
        "instruccion": "Crea una tupla 'coordenadas' con valores 5,-3,2.5 e imprime.",
        "codigo": "coordenadas=(5,-3,2.5)\nprint(coordenadas)"
    },
    {
        "numero": 9,
        "tema": "Variable tipo diccionario",
        "instruccion": "Crea un diccionario 'empleado' con nombre, edad, salario, departamento e imprime.",
        "codigo": "empleado={'nombre':'Carlos','edad':30,'salario':45000,'departamento':'TI'}\nprint(empleado)"
    },
    {
        "numero": 10,
        "tema": "Función type()",
        "instruccion": "Transforma 'precio' a flotante, usa type() para guardar el tipo en 'clase' e imprime.",
        "codigo": "precio='99.99'\nprecio=float(precio)\nclase=type(precio)\nprint(precio,clase)"
    },
    {
        "numero": 11,
        "tema": "Operaciones aritméticas",
        "instruccion": "Guarda el resultado de restar Y de X en 'diferencia' e imprime.",
        "codigo": "X=50\nY=15\ndiferencia=X-Y\nprint(diferencia)"
    },
    {
        "numero": 12,
        "tema": "Asignación compuesta",
        "instruccion": "Resta 5 a 'saldo' usando operador de asignación compuesta e imprime.",
        "codigo": "saldo=100\nsaldo-=5\nprint(saldo)"
    },
    {
        "numero": 13,
        "tema": "Operador booleano",
        "instruccion": "Modifica el operador booleano para que imprima 'False'.",
        "codigo": "validacion=False or True\nprint(validacion)"
    },
    {
        "numero": 14,
        "tema": "Operador relacional",
        "instruccion": "Completa el código con un operador relacional e imprime el resultado.",
        "codigo": "a=8\nb=12\nresultado=a<b\nprint(resultado)"
    },
    {
        "numero": 15,
        "tema": "Métodos de cadena",
        "instruccion": "Transforma 'programacion' a mayúsculas e imprime.",
        "codigo": "palabra='programacion'\nprint(palabra.upper())"
    },
    {
        "numero": 16,
        "tema": "Método append",
        "instruccion": "Añade 'mango' a la lista 'frutas' con append() e imprime.",
        "codigo": "frutas=['manzana','banana','naranja']\nfrutas.append('mango')\nprint(frutas)"
    },
    {
        "numero": 17,
        "tema": "Método update",
        "instruccion": "Crea diccionario 'calificaciones_nuevas' y úsalo para actualizar 'calificaciones'.",
        "codigo": "calificaciones={'Pedro':85,'María':90}\ncalificaciones_nuevas={'Laura':95,'Fernando':88,'Rosa':92}\ncalificaciones.update(calificaciones_nuevas)\nprint(calificaciones)"
    },
    {
        "numero": 18,
        "tema": "Condicional if",
        "instruccion": "Verifica si 'edad' es mayor o igual a 18 e imprime con un mensaje.",
        "codigo": "edad=22\nif edad>=18:\n    print(edad,'Es mayor de edad:')"
    },
    {
        "numero": 19,
        "tema": "Condicional if-else",
        "instruccion": "Determina si un estudiante aprobó (>=60) o reprobó con if-else.",
        "codigo": "calificacion=75\nif calificacion>=60:\n    print('Estudiante aprobado')\nelse:\n    print('Estudiante reprobado')"
    },
    {
        "numero": 20,
        "tema": "Condicional elif",
        "instruccion": "Clasifica un atleta por su tiempo usando if-elif-else.",
        "codigo": "tiempo=12.5\nif tiempo<11:\n    print('Excelente')\nelif tiempo<13:\n    print('Bueno')\nelse:\n    print('Necesita mejorar')"
    },
    {
        "numero": 21,
        "tema": "Condicional anidado",
        "instruccion": "Verifica si 'nivel' es 'admin' para acceso total o restringido.",
        "codigo": "nivel='usuario'\nif nivel=='admin':\n    print('Acceso total')\nelse:\n    print('Acceso restringido')"
    },
    {
        "numero": 22,
        "tema": "Ciclo while",
        "instruccion": "Crea ciclo while que pare cuando contador llegue a 50.",
        "codigo": "contador=0\nwhile contador<50:\n    print(contador)\n    contador+=10"
    },
    {
        "numero": 23,
        "tema": "Ciclo for",
        "instruccion": "Lee una palabra con for, agrega cada carácter a lista con append().",
        "codigo": "palabra='Python'\ncaracteres=[]\nfor carácter in palabra:\n    caracteres.append(carácter)\nprint(caracteres)"
    },
    {
        "numero": 24,
        "tema": "Ciclos anidados",
        "instruccion": "Obtiene elementos de matriz con ciclos for anidados.",
        "codigo": "matriz=[[1,2,3],[4,5,6],[7,8,9]]\nfor fila in matriz:\n    for valor in fila:\n        print(valor)"
    },
    {
        "numero": 25,
        "tema": "Función con return",
        "instruccion": "Crea función 'restar' que calcule resta de dos números con return.",
        "codigo": "def restar(a,b):\n    resultado=a-b\n    return resultado\nprint(restar(20,8))"
    },
    {
        "numero": 26,
        "tema": "Try-except",
        "instruccion": "Usa try-except para manejar errores de conversión de tipos.",
        "codigo": "try:\n    numero=int('50')\n    resultado=numero*5\n    print(resultado)\nexcept ValueError:\n    print('Error: Valor inválido')"
    },
    {
        "numero": 27,
        "tema": "Lectura de archivos",
        "instruccion": "Lee archivo 'el_principito.txt' con with en modo lectura UTF-8.",
        "codigo": "# with open('el_principito.txt','r',encoding='UTF-8') as archivo:\n#     for línea in archivo:\n#         print(línea)"
    },
    {
        "numero": 28,
        "tema": "Librería random",
        "instruccion": "Genera número aleatorio entre 10 y 50 con random.randint().",
        "codigo": "import random\nnumero_aleatorio=random.randint(10,50)\nprint(numero_aleatorio)"
    }
]

# Agregar datos a la hoja
for i, ejercicio in enumerate(ejercicios, start=2):
    ws[f'A{i}'] = ejercicio['numero']
    ws[f'B{i}'] = ejercicio['tema']
    ws[f'C{i}'] = ejercicio['instruccion']
    ws[f'D{i}'] = ejercicio['codigo']
    
    # Aplicar bordes y alineación
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{i}']
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 50

# Ajustar alto de filas para mejor legibilidad
for row in ws.iter_rows(min_row=2, max_row=len(ejercicios)+1):
    ws.row_dimensions[row[0].row].height = 60

# Guardar el archivo
wb.save('Respuestas_Ejercicios_Python.xlsx')
print("✓ Archivo 'Respuestas_Ejercicios_Python.xlsx' creado exitosamente")
